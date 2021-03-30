import openpyxl as xl
import itertools
import json
import os

SOURCE_UNITS_TABLE = os.path.join(os.path.dirname(__file__), 'Units.xlsx')
EXPORT_UNITS_SYSTEM = os.path.join(os.path.dirname(__file__), 'UnitsSystem.json')


def load_table(path, sheet, left_top_col, left_top_row):
    '''
    Read one table from Excel file. The table starts with one line header.
    path - file path
    sheet - sheet (tab) name in the workbook
    Coordinates of the table's left top cell:
    left_top_col - column
    left_top_row - row
    '''
    wb = xl.load_workbook(path)
    ws = wb[sheet]
    header = []
    for c in itertools.count(left_top_col):
        col_name = ws.cell(left_top_row, c).value
        if not col_name:    # table header ends with first empty cell
            break
        header.append(col_name)
    data = []
    for row_data in ws.iter_rows(min_row=left_top_row + 1, min_col=left_top_col, max_col=left_top_col + len(header) - 1, values_only=True):
        if next(filter(bool, row_data), None) is None:  # table ends with first empty line
            break
        data.append(row_data)
    return header, data


header, rows = load_table(SOURCE_UNITS_TABLE, 'Units', 1, 1)
units = {}
# unit_value = base_unit_value * k + b
for row in rows:
    line = dict(zip(header, row))
    if line['Drop'] != 'Yes':
        base_unit = line['BaseUnit'].strip()
        unit = line['Unit'].strip()
        dimention = line['Dimension Name'].strip()
        unit_info = {'k': float(line['Scale'])}
        b = line['Offset']
        if b is not None:
            b = float(b)
            if b != 0:
                unit_info['b'] = b
        compat_units = units.setdefault(dimention, {})
        if base_unit not in compat_units:
            compat_units[base_unit] = {'k': 1.0, 'base': True}
        if unit != base_unit:
            compat_units[unit] = unit_info
# find case-sensitive units
delete_later = set()
for dimention, compat_units in units.items():
    for unit in compat_units:
        lunit = unit.lower()
        for dimention2, compat_units2 in units.items():
            for unit2 in compat_units2:
                lunit2 = unit2.lower()
                if dimention == dimention2 and unit == unit2:   # it's me, skip
                    continue
                assert unit != unit2, f'Units must be unique. Double entry of {dimention}."{unit}" and {dimention2}."{unit2}"'
                if lunit == lunit2:
                    if dimention == dimention2:     # same dimention, compatible units with different case
                        unit_info = compat_units[unit]
                        unit2_info = compat_units2[unit2]
                        if unit_info['k'] == unit2_info['k'] and unit_info.get('b') == unit2_info.get('b'):  # same unit with different case
                            if (dimention, unit) not in delete_later and (dimention2, unit2) not in delete_later:
                                if unit_info.get('base', False):    # not going to delete a base unit
                                    delete_later.add((dimention2, unit2))
                                else:
                                    delete_later.add((dimention, unit))
                        else:   # different units with case difference
                            compat_units[unit]['case-sensitive'] = True
                    else:   # units from diferent dimentions are always different
                        compat_units[unit]['case-sensitive'] = True

for dimention, unit in delete_later:
    del units[dimention][unit]

with open(EXPORT_UNITS_SYSTEM, 'w') as f:
    json.dump(units, f, sort_keys=True, indent='\t')
