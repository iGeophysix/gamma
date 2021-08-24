# Log family properties database
# Manually run this script to update FamilyProperties in Redis

import itertools
import os
from typing import Tuple

import openpyxl as xl

from components.database.RedisStorage import RedisStorage, FAMILY_PROPERTIES_TABLE_NAME
from settings import BASE_DIR

SOURCE_FAMSET_BOOK = os.path.join(BASE_DIR,
                                  'components',
                                  'importexport',
                                  'rules',
                                  'src',
                                  'FamilyProperties.xlsx')

DEFAULT_THICKNESS = 1
DEFAULT_COLOR = '000000'  # black in hex RGB24


def capitalize(s: str):
    '''
    Make the first letter of all words capital.
    '''
    return ' '.join(map(str.capitalize, s.split(' ')))


def read_cell_background_color(header, row_data) -> Tuple[int]:
    '''
    TODO: Add comment
    '''

    color = DEFAULT_COLOR
    cell_background = row_data[header['Color']].fill
    if cell_background is not None and cell_background.patternType == 'solid':
        bg_color = cell_background.start_color
        # theme color is not supported
        if bg_color.type == 'rgb':
            color = bg_color.rgb[-6:]  # cut off alpha channel
        elif bg_color.type == 'indexed':
            color = xl.styles.colors.COLOR_INDEX[bg_color.indexed][-6:]
    else:  # color may be typed as cell text
        s_color = row_data[header['Color']].value
        if s_color is not None:
            color = s_color[-6:]

    # converts 0A0B0C to (10, 11, 12)
    rgb_triplet = tuple(int(color[n: n + 2], 16) for n in range(0, 6, 2))

    return rgb_triplet


def parse_excel_table(src_path, sheet) -> dict:
    '''
    Generates family properties dictionary from excel sheet
    '''
    left_top_row = 1  # table start
    left_top_col = 1  #

    wb = xl.load_workbook(src_path, read_only=True, data_only=True)
    ws = wb[sheet]
    # read header
    header = {}  # maping column name to column index
    for c in itertools.count(left_top_col):
        col_name = ws.cell(left_top_row, c).value
        if not col_name:  # table header ends with first empty cell
            break
        header[col_name] = c - 1
    # read table data
    data = {}

    iter_rows = ws.iter_rows(min_row=left_top_row + 1,
                             min_col=left_top_col,
                             max_col=left_top_col + len(header) - 1)

    for row_data in iter_rows:
        family = row_data[header['Family']].value
        if not family:
            break
        family = capitalize(family)
        splice = row_data[header['Splice']].value == 'TRUE'
        mnemonic = row_data[header['Mnemonic']].value or family
        min_ = row_data[header['Min']].value
        max_ = row_data[header['Max']].value
        unit = row_data[header['Unit']].value
        thickness = row_data[header['Thickness']].value or DEFAULT_THICKNESS

        # read cell backgroud color
        rgb_triplet = read_cell_background_color(header, row_data)

        data[family] = {'splice': splice,
                        'mnemonic': mnemonic,
                        'min': min_,
                        'max': max_,
                        'unit': unit,
                        'color': rgb_triplet,
                        'thickness': thickness}

    return data


class FamilyProperties:
    '''
    Family properties storage:
    {
     'Family': {'splice': bool,
                'mnemonic': str,
                'min': float or None,
                'max': float or None,
                'unit': str or None,
                'color': (R, G, B),
                'thickness': int}
    }
    '''
    def __init__(self):
        self._s = RedisStorage()
        assert self._s.object_exists(FAMILY_PROPERTIES_TABLE_NAME), 'Common data is absent'

    @staticmethod
    def build_family_properties():
        data = parse_excel_table(SOURCE_FAMSET_BOOK, 'FamilyProperties')
        s = RedisStorage()
        s.object_delete(FAMILY_PROPERTIES_TABLE_NAME)
        s.table_key_set(FAMILY_PROPERTIES_TABLE_NAME, mapping=data)

    def exists(self, item) -> bool:
        return self._s.table_key_exists(FAMILY_PROPERTIES_TABLE_NAME, item)

    def __getitem__(self, item) -> dict:
        if self._s.table_key_exists(FAMILY_PROPERTIES_TABLE_NAME, item):
            return self._s.table_key_get(FAMILY_PROPERTIES_TABLE_NAME, item)
        else:
            return {}

    def __setitem__(self, key, value):
        self._s.table_key_set(FAMILY_PROPERTIES_TABLE_NAME, key, value)

    def items(self):
        for key in self._s.table_keys(FAMILY_PROPERTIES_TABLE_NAME):
            yield (key, self._s.table_key_get(FAMILY_PROPERTIES_TABLE_NAME, key))


if __name__ == '__main__':
    # Export Excel sheet to JSON family properties file
    FamilyProperties.build_family_properties()
